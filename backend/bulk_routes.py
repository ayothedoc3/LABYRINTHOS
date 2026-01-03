"""
Bulk Upload Routes for Labyrinth Operating System
Supports CSV, JSON, and Excel file uploads with preview and validation
"""

from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any, Optional
from datetime import datetime, timezone
from pydantic import BaseModel, ValidationError
import json
import csv
import io
import uuid

# Import models
from models import (
    FunctionType, LevelType, ClientPackage,
    PlaybookCreate, Playbook,
    SOPCreate, SOP, SOPStep,
    TalentCreate, Talent, CompetencyScores,
    ContractCreate, Contract, ContractBoundary,
    KPICreate, KPI, KPIThresholds
)

bulk_router = APIRouter(prefix="/bulk", tags=["Bulk Upload"])

# Global db reference - set from server.py
db = None

def set_db(database):
    global db
    db = database


# ==================== RESPONSE MODELS ====================

class ValidationError_Item(BaseModel):
    row: int
    field: str
    error: str
    value: Any = None

class PreviewRow(BaseModel):
    row_number: int
    data: Dict[str, Any]
    is_valid: bool
    errors: List[ValidationError_Item] = []

class BulkUploadPreview(BaseModel):
    entity_type: str
    total_rows: int
    valid_rows: int
    invalid_rows: int
    preview_data: List[PreviewRow]
    columns: List[str]

class BulkImportResult(BaseModel):
    entity_type: str
    total_processed: int
    successful: int
    failed: int
    errors: List[Dict[str, Any]] = []


# ==================== TEMPLATE DEFINITIONS ====================

PLAYBOOK_COLUMNS = ["playbook_id", "name", "function", "level", "min_tier", "description", "linked_sop_ids"]
SOP_COLUMNS = ["sop_id", "name", "function", "linked_playbook_ids", "template_required", "estimated_time_minutes"]
TALENT_COLUMNS = ["name", "email", "function", "subfunction", "communication", "technical_skills", "problem_solving", "time_management", "leadership", "adaptability", "domain_expertise", "tags", "notes"]
CONTRACT_COLUMNS = ["talent_id", "client_name", "client_package", "assigned_playbook_ids", "start_date", "hourly_rate", "monthly_retainer"]
KPI_COLUMNS = ["kpi_id", "name", "function", "description", "unit", "target", "yellow_threshold", "red_threshold", "is_higher_better", "measurement_formula"]

SAMPLE_DATA = {
    "playbooks": [
        {
            "playbook_id": "SALES-ACQ-01",
            "name": "New Client Acquisition",
            "function": "SALES",
            "level": "ACQUIRE",
            "min_tier": 2,
            "description": "Standard playbook for acquiring new clients through cold outreach and referrals",
            "linked_sop_ids": "SOP-SALES-001,SOP-SALES-002"
        },
        {
            "playbook_id": "MKTG-MAIN-01",
            "name": "Content Marketing Maintenance",
            "function": "MARKETING",
            "level": "MAINTAIN",
            "min_tier": 1,
            "description": "Ongoing content creation and distribution playbook",
            "linked_sop_ids": "SOP-MKTG-001"
        }
    ],
    "sops": [
        {
            "sop_id": "SOP-SALES-001",
            "name": "Lead Qualification Process",
            "function": "SALES",
            "linked_playbook_ids": "SALES-ACQ-01",
            "template_required": "Lead Scoring Template",
            "estimated_time_minutes": 30
        },
        {
            "sop_id": "SOP-MKTG-001",
            "name": "Blog Post Creation",
            "function": "MARKETING",
            "linked_playbook_ids": "MKTG-MAIN-01",
            "template_required": "Content Brief Template",
            "estimated_time_minutes": 60
        }
    ],
    "talents": [
        {
            "name": "John Smith",
            "email": "john.smith@example.com",
            "function": "SALES",
            "subfunction": "Business Development",
            "communication": 4.5,
            "technical_skills": 3.5,
            "problem_solving": 4.0,
            "time_management": 4.0,
            "leadership": 3.5,
            "adaptability": 4.0,
            "domain_expertise": 4.5,
            "tags": "senior,closer,enterprise",
            "notes": "Top performer in Q3"
        },
        {
            "name": "Jane Doe",
            "email": "jane.doe@example.com",
            "function": "MARKETING",
            "subfunction": "Content",
            "communication": 5.0,
            "technical_skills": 3.0,
            "problem_solving": 4.0,
            "time_management": 4.5,
            "leadership": 3.0,
            "adaptability": 4.5,
            "domain_expertise": 4.0,
            "tags": "writer,creative",
            "notes": "Excellent content creator"
        }
    ],
    "contracts": [
        {
            "talent_id": "existing_talent_id_here",
            "client_name": "Acme Corporation",
            "client_package": "SILVER",
            "assigned_playbook_ids": "SALES-ACQ-01",
            "start_date": "2025-01-01",
            "hourly_rate": 75.00,
            "monthly_retainer": 5000.00
        }
    ],
    "kpis": [
        {
            "kpi_id": "KPI-SALES-001",
            "name": "Lead Conversion Rate",
            "function": "SALES",
            "description": "Percentage of qualified leads converted to customers",
            "unit": "%",
            "target": 25.0,
            "yellow_threshold": 15.0,
            "red_threshold": 10.0,
            "is_higher_better": True,
            "measurement_formula": "(Conversions / Qualified Leads) * 100"
        },
        {
            "kpi_id": "KPI-MKTG-001",
            "name": "Content Engagement Rate",
            "function": "MARKETING",
            "description": "Average engagement rate across content pieces",
            "unit": "%",
            "target": 5.0,
            "yellow_threshold": 3.0,
            "red_threshold": 1.5,
            "is_higher_better": True,
            "measurement_formula": "(Engagements / Impressions) * 100"
        }
    ]
}


# ==================== HELPER FUNCTIONS ====================

def parse_list_field(value: str) -> List[str]:
    """Parse comma-separated string into list"""
    if not value or value == "":
        return []
    return [v.strip() for v in str(value).split(",") if v.strip()]


def parse_bool_field(value: Any) -> bool:
    """Parse boolean from various formats"""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ("true", "yes", "1", "t", "y")
    return bool(value)


def parse_float_field(value: Any, default: float = 0.0) -> float:
    """Parse float safely"""
    try:
        return float(value) if value is not None and value != "" else default
    except (ValueError, TypeError):
        return default


def parse_int_field(value: Any, default: int = 0) -> int:
    """Parse int safely"""
    try:
        return int(float(value)) if value is not None and value != "" else default
    except (ValueError, TypeError):
        return default


async def parse_uploaded_file(file: UploadFile) -> tuple[List[Dict[str, Any]], List[str]]:
    """Parse uploaded file (CSV, JSON, or Excel) and return rows and columns"""
    content = await file.read()
    filename = file.filename.lower()
    
    if filename.endswith('.json'):
        data = json.loads(content.decode('utf-8'))
        if isinstance(data, list) and len(data) > 0:
            columns = list(data[0].keys())
            return data, columns
        raise HTTPException(status_code=400, detail="JSON file must contain an array of objects")
    
    elif filename.endswith('.csv'):
        text_content = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text_content))
        rows = list(reader)
        columns = reader.fieldnames or []
        return rows, columns
    
    elif filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if not header:
                raise HTTPException(status_code=400, detail="Excel file has no header row")
            
            columns = [str(h) if h else f"col_{i}" for i, h in enumerate(header)]
            rows = []
            for row in rows_iter:
                if any(cell is not None for cell in row):
                    row_dict = {columns[i]: (row[i] if i < len(row) else None) for i in range(len(columns))}
                    rows.append(row_dict)
            
            return rows, columns
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")
    
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use CSV, JSON, or Excel (.xlsx)")


def validate_playbook_row(row: Dict[str, Any], row_num: int) -> tuple[Optional[Dict], List[ValidationError_Item]]:
    """Validate a single playbook row"""
    errors = []
    
    # Required fields
    if not row.get("playbook_id"):
        errors.append(ValidationError_Item(row=row_num, field="playbook_id", error="Required field missing"))
    if not row.get("name"):
        errors.append(ValidationError_Item(row=row_num, field="name", error="Required field missing"))
    
    # Validate function enum
    func_value = str(row.get("function", "")).upper()
    if func_value not in [f.value for f in FunctionType]:
        errors.append(ValidationError_Item(row=row_num, field="function", error=f"Invalid value. Must be one of: {[f.value for f in FunctionType]}", value=row.get("function")))
    
    # Validate level enum
    level_value = str(row.get("level", "")).upper()
    if level_value not in [l.value for l in LevelType]:
        errors.append(ValidationError_Item(row=row_num, field="level", error=f"Invalid value. Must be one of: {[l.value for l in LevelType]}", value=row.get("level")))
    
    # Validate min_tier
    min_tier = parse_int_field(row.get("min_tier"), 1)
    if min_tier < 1 or min_tier > 3:
        errors.append(ValidationError_Item(row=row_num, field="min_tier", error="Must be between 1 and 3", value=row.get("min_tier")))
    
    if errors:
        return None, errors
    
    # Create validated data
    validated = {
        "playbook_id": str(row["playbook_id"]).strip(),
        "name": str(row["name"]).strip(),
        "function": func_value,
        "level": level_value,
        "min_tier": min_tier,
        "description": str(row.get("description", "")).strip(),
        "linked_sop_ids": parse_list_field(row.get("linked_sop_ids", ""))
    }
    
    return validated, errors


def validate_sop_row(row: Dict[str, Any], row_num: int) -> tuple[Optional[Dict], List[ValidationError_Item]]:
    """Validate a single SOP row"""
    errors = []
    
    if not row.get("sop_id"):
        errors.append(ValidationError_Item(row=row_num, field="sop_id", error="Required field missing"))
    if not row.get("name"):
        errors.append(ValidationError_Item(row=row_num, field="name", error="Required field missing"))
    
    func_value = str(row.get("function", "")).upper()
    if func_value not in [f.value for f in FunctionType]:
        errors.append(ValidationError_Item(row=row_num, field="function", error=f"Invalid value. Must be one of: {[f.value for f in FunctionType]}", value=row.get("function")))
    
    if errors:
        return None, errors
    
    validated = {
        "sop_id": str(row["sop_id"]).strip(),
        "name": str(row["name"]).strip(),
        "function": func_value,
        "linked_playbook_ids": parse_list_field(row.get("linked_playbook_ids", "")),
        "template_required": str(row.get("template_required", "")).strip(),
        "steps": [],
        "estimated_time_minutes": parse_int_field(row.get("estimated_time_minutes"), 30)
    }
    
    return validated, errors


def validate_talent_row(row: Dict[str, Any], row_num: int) -> tuple[Optional[Dict], List[ValidationError_Item]]:
    """Validate a single talent row"""
    errors = []
    
    if not row.get("name"):
        errors.append(ValidationError_Item(row=row_num, field="name", error="Required field missing"))
    if not row.get("email"):
        errors.append(ValidationError_Item(row=row_num, field="email", error="Required field missing"))
    
    func_value = str(row.get("function", "")).upper()
    if func_value not in [f.value for f in FunctionType]:
        errors.append(ValidationError_Item(row=row_num, field="function", error=f"Invalid value. Must be one of: {[f.value for f in FunctionType]}", value=row.get("function")))
    
    # Validate competency scores (1-5)
    score_fields = ["communication", "technical_skills", "problem_solving", "time_management", "leadership", "adaptability", "domain_expertise"]
    for field in score_fields:
        val = parse_float_field(row.get(field), 3.0)
        if val < 1.0 or val > 5.0:
            errors.append(ValidationError_Item(row=row_num, field=field, error="Score must be between 1.0 and 5.0", value=row.get(field)))
    
    if errors:
        return None, errors
    
    validated = {
        "name": str(row["name"]).strip(),
        "email": str(row["email"]).strip(),
        "function": func_value,
        "subfunction": str(row.get("subfunction", "")).strip() if row.get("subfunction") else None,
        "competency_scores": {
            "communication": parse_float_field(row.get("communication"), 3.0),
            "technical_skills": parse_float_field(row.get("technical_skills"), 3.0),
            "problem_solving": parse_float_field(row.get("problem_solving"), 3.0),
            "time_management": parse_float_field(row.get("time_management"), 3.0),
            "leadership": parse_float_field(row.get("leadership"), 3.0),
            "adaptability": parse_float_field(row.get("adaptability"), 3.0),
            "domain_expertise": parse_float_field(row.get("domain_expertise"), 3.0)
        },
        "tags": parse_list_field(row.get("tags", "")),
        "notes": str(row.get("notes", "")).strip() if row.get("notes") else None
    }
    
    return validated, errors


def validate_contract_row(row: Dict[str, Any], row_num: int) -> tuple[Optional[Dict], List[ValidationError_Item]]:
    """Validate a single contract row"""
    errors = []
    
    if not row.get("talent_id"):
        errors.append(ValidationError_Item(row=row_num, field="talent_id", error="Required field missing"))
    if not row.get("client_name"):
        errors.append(ValidationError_Item(row=row_num, field="client_name", error="Required field missing"))
    
    package_value = str(row.get("client_package", "")).upper()
    if package_value not in [p.value for p in ClientPackage]:
        errors.append(ValidationError_Item(row=row_num, field="client_package", error=f"Invalid value. Must be one of: {[p.value for p in ClientPackage]}", value=row.get("client_package")))
    
    if errors:
        return None, errors
    
    # Parse start_date
    start_date_str = str(row.get("start_date", ""))
    try:
        if start_date_str:
            start_date = datetime.fromisoformat(start_date_str.replace("Z", "+00:00"))
        else:
            start_date = datetime.now(timezone.utc)
    except ValueError:
        start_date = datetime.now(timezone.utc)
    
    validated = {
        "talent_id": str(row["talent_id"]).strip(),
        "client_name": str(row["client_name"]).strip(),
        "client_package": package_value,
        "assigned_playbook_ids": parse_list_field(row.get("assigned_playbook_ids", "")),
        "assigned_sop_ids": [],
        "kpi_ids": [],
        "boundaries": {
            "max_hours_per_week": 40,
            "response_time_hours": 24,
            "deliverable_quality_min": 3.5,
            "escalation_threshold_days": 3
        },
        "start_date": start_date.isoformat(),
        "hourly_rate": parse_float_field(row.get("hourly_rate")),
        "monthly_retainer": parse_float_field(row.get("monthly_retainer"))
    }
    
    return validated, errors


def validate_kpi_row(row: Dict[str, Any], row_num: int) -> tuple[Optional[Dict], List[ValidationError_Item]]:
    """Validate a single KPI row"""
    errors = []
    
    if not row.get("kpi_id"):
        errors.append(ValidationError_Item(row=row_num, field="kpi_id", error="Required field missing"))
    if not row.get("name"):
        errors.append(ValidationError_Item(row=row_num, field="name", error="Required field missing"))
    
    func_value = str(row.get("function", "")).upper()
    if func_value not in [f.value for f in FunctionType]:
        errors.append(ValidationError_Item(row=row_num, field="function", error=f"Invalid value. Must be one of: {[f.value for f in FunctionType]}", value=row.get("function")))
    
    if errors:
        return None, errors
    
    validated = {
        "kpi_id": str(row["kpi_id"]).strip(),
        "name": str(row["name"]).strip(),
        "function": func_value,
        "description": str(row.get("description", "")).strip(),
        "unit": str(row.get("unit", "%")).strip(),
        "thresholds": {
            "target": parse_float_field(row.get("target"), 0),
            "yellow_threshold": parse_float_field(row.get("yellow_threshold"), 0),
            "red_threshold": parse_float_field(row.get("red_threshold"), 0),
            "is_higher_better": parse_bool_field(row.get("is_higher_better", True))
        },
        "measurement_formula": str(row.get("measurement_formula", "")).strip()
    }
    
    return validated, errors


VALIDATORS = {
    "playbooks": validate_playbook_row,
    "sops": validate_sop_row,
    "talents": validate_talent_row,
    "contracts": validate_contract_row,
    "kpis": validate_kpi_row
}


# ==================== TEMPLATE DOWNLOAD ENDPOINTS ====================

@bulk_router.get("/template/{entity_type}")
async def download_template(entity_type: str, format: str = "csv"):
    """Download a template file with sample data"""
    if entity_type not in SAMPLE_DATA:
        raise HTTPException(status_code=400, detail=f"Invalid entity type. Must be one of: {list(SAMPLE_DATA.keys())}")
    
    sample = SAMPLE_DATA[entity_type]
    
    if format == "json":
        content = json.dumps(sample, indent=2)
        media_type = "application/json"
        filename = f"{entity_type}_template.json"
    elif format == "csv":
        output = io.StringIO()
        if sample:
            writer = csv.DictWriter(output, fieldnames=sample[0].keys())
            writer.writeheader()
            writer.writerows(sample)
        content = output.getvalue()
        media_type = "text/csv"
        filename = f"{entity_type}_template.csv"
    elif format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = entity_type.capitalize()
            
            if sample:
                # Header
                headers = list(sample[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                # Data rows
                for row_idx, row_data in enumerate(sample, 2):
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col, value=row_data.get(header, ""))
                
                # Adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column].width = min(max_length + 2, 50)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            content = output.getvalue()
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"{entity_type}_template.xlsx"
            
            return StreamingResponse(
                io.BytesIO(content),
                media_type=media_type,
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel support not available")
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'csv', 'json', or 'xlsx'")
    
    return StreamingResponse(
        io.BytesIO(content.encode() if isinstance(content, str) else content),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== PREVIEW ENDPOINT ====================

@bulk_router.post("/preview/{entity_type}", response_model=BulkUploadPreview)
async def preview_upload(entity_type: str, file: UploadFile = File(...)):
    """Preview and validate uploaded file before import"""
    if entity_type not in VALIDATORS:
        raise HTTPException(status_code=400, detail=f"Invalid entity type. Must be one of: {list(VALIDATORS.keys())}")
    
    rows, columns = await parse_uploaded_file(file)
    validator = VALIDATORS[entity_type]
    
    preview_data = []
    valid_count = 0
    invalid_count = 0
    
    for idx, row in enumerate(rows):
        row_num = idx + 2  # Account for header row
        validated, errors = validator(row, row_num)
        
        is_valid = len(errors) == 0
        if is_valid:
            valid_count += 1
        else:
            invalid_count += 1
        
        preview_data.append(PreviewRow(
            row_number=row_num,
            data=row,
            is_valid=is_valid,
            errors=errors
        ))
    
    return BulkUploadPreview(
        entity_type=entity_type,
        total_rows=len(rows),
        valid_rows=valid_count,
        invalid_rows=invalid_count,
        preview_data=preview_data,
        columns=columns
    )


# ==================== IMPORT ENDPOINT ====================

@bulk_router.post("/import/{entity_type}", response_model=BulkImportResult)
async def import_data(entity_type: str, file: UploadFile = File(...), skip_invalid: bool = True):
    """Import validated data from uploaded file"""
    if entity_type not in VALIDATORS:
        raise HTTPException(status_code=400, detail=f"Invalid entity type. Must be one of: {list(VALIDATORS.keys())}")
    
    rows, _ = await parse_uploaded_file(file)
    validator = VALIDATORS[entity_type]
    
    successful = 0
    failed = 0
    errors = []
    
    # Collection mapping
    collections = {
        "playbooks": db.playbooks,
        "sops": db.sops,
        "talents": db.talents,
        "contracts": db.contracts,
        "kpis": db.kpis
    }
    
    # ID field mapping
    id_fields = {
        "playbooks": "playbook_id",
        "sops": "sop_id",
        "talents": None,  # Uses email for duplicate check
        "contracts": None,  # No duplicate check
        "kpis": "kpi_id"
    }
    
    collection = collections[entity_type]
    id_field = id_fields[entity_type]
    
    for idx, row in enumerate(rows):
        row_num = idx + 2
        validated, validation_errors = validator(row, row_num)
        
        if validation_errors:
            if skip_invalid:
                failed += 1
                errors.append({
                    "row": row_num,
                    "errors": [e.model_dump() for e in validation_errors]
                })
                continue
            else:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Validation failed at row {row_num}: {validation_errors[0].error}"
                )
        
        try:
            # Check for duplicates
            if id_field and validated.get(id_field):
                existing = await collection.find_one({id_field: validated[id_field]})
                if existing:
                    # Update existing record
                    validated["updated_at"] = datetime.now(timezone.utc).isoformat()
                    await collection.update_one(
                        {id_field: validated[id_field]},
                        {"$set": validated}
                    )
                    successful += 1
                    continue
            elif entity_type == "talents":
                # Check by email for talents
                existing = await collection.find_one({"email": validated["email"]})
                if existing:
                    validated["updated_at"] = datetime.now(timezone.utc).isoformat()
                    await collection.update_one(
                        {"email": validated["email"]},
                        {"$set": validated}
                    )
                    successful += 1
                    continue
            
            # Insert new record
            validated["id"] = str(uuid.uuid4())
            validated["created_at"] = datetime.now(timezone.utc).isoformat()
            validated["updated_at"] = datetime.now(timezone.utc).isoformat()
            validated["is_active"] = True
            
            # Calculate tier for talents
            if entity_type == "talents":
                scores = validated.get("competency_scores", {})
                avg = sum(scores.values()) / len(scores) if scores else 3.0
                validated["tier_score"] = round(avg, 2)
                if avg <= 2.0:
                    validated["current_tier"] = 1
                elif avg <= 3.5:
                    validated["current_tier"] = 2
                else:
                    validated["current_tier"] = 3
            
            await collection.insert_one(validated)
            successful += 1
            
        except Exception as e:
            failed += 1
            errors.append({
                "row": row_num,
                "error": str(e)
            })
    
    return BulkImportResult(
        entity_type=entity_type,
        total_processed=len(rows),
        successful=successful,
        failed=failed,
        errors=errors
    )


# ==================== GET AVAILABLE TEMPLATES INFO ====================

@bulk_router.get("/templates-info")
async def get_templates_info():
    """Get information about available templates"""
    return {
        "available_templates": [
            {
                "entity_type": "playbooks",
                "description": "Business playbooks with tier and function assignments",
                "columns": PLAYBOOK_COLUMNS,
                "required_columns": ["playbook_id", "name", "function", "level"],
                "formats": ["csv", "json", "xlsx"]
            },
            {
                "entity_type": "sops",
                "description": "Standard Operating Procedures linked to playbooks",
                "columns": SOP_COLUMNS,
                "required_columns": ["sop_id", "name", "function"],
                "formats": ["csv", "json", "xlsx"]
            },
            {
                "entity_type": "talents",
                "description": "Team members with competency scores",
                "columns": TALENT_COLUMNS,
                "required_columns": ["name", "email", "function"],
                "formats": ["csv", "json", "xlsx"]
            },
            {
                "entity_type": "contracts",
                "description": "Client contracts linked to talents",
                "columns": CONTRACT_COLUMNS,
                "required_columns": ["talent_id", "client_name", "client_package"],
                "formats": ["csv", "json", "xlsx"]
            },
            {
                "entity_type": "kpis",
                "description": "Key Performance Indicators with thresholds",
                "columns": KPI_COLUMNS,
                "required_columns": ["kpi_id", "name", "function"],
                "formats": ["csv", "json", "xlsx"]
            }
        ],
        "supported_formats": ["csv", "json", "xlsx"],
        "valid_functions": [f.value for f in FunctionType],
        "valid_levels": [l.value for l in LevelType],
        "valid_packages": [p.value for p in ClientPackage]
    }
